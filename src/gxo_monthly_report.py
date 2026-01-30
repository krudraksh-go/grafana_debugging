import math
import json
import smtplib
import ssl
from collections import OrderedDict
from datetime import date, datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import airflow
import pandas as pd
from airflow import DAG
from airflow.operators.python import PythonOperator
from jinja2 import Environment
from openpyxl import Workbook, load_workbook

from utils.CommonFunction import InfluxData, CommonFunction
from influxdb import InfluxDBClient
from config import (
    SMTP_USERNAME,
    SMTP_PASSWORD,
    SMTP_HOST,
    SODIMAC_SENDER_EMAIL,
    SODIMAC_RECEIVER_EMAIL,
)

SITES = [{
    "id": "adidas",
    "name": "Adidas",
    "file": "dags/files/gxo_monthly_report_adidas.xlsx"
},
# {
#     "id": "apple",
#     "name": "Apple",
#     "file": "dags/files/gxo_monthly_report_apple.xlsx"
# }
]

MAIL_TEMPLATE = """
{% for site in sites_data_list %}
    <h3>{{ site.name }}</h3>
    <table style="width: 800px; font-size: 18px; border: solid #f8f9fa;">
        <thead style="background: antiquewhite;">
            <th style="text-align: left;">Name</th>
            <th style="text-align: center;">1st to 7th</th>
            <th style="text-align: center;">8th to 14th</th>
            <th style="text-align: center;">15th to 21st</th>
            <th style="text-align: center;">22nd to remaining</th>
            <th style="text-align: center;">Monthly_Avg</th>
        </thead>
        <tbody style="background: floralwhite;">
            {% for row in site.rows %}
                <tr>
                    <td style="text-align: left;">{{ row.name }}</td>
                    {% for value in row.data_list %}
                        <td style="text-align: center;">{{ value }}</td>
                    {% endfor %}
                </tr>
            {% endfor %}
        </tbody>
    </table>
{% endfor %}
"""


KPI_NAMES = [
    "Avg Audit Completed",
    "Avg Audit Not Started",
    "Avg Audit Unprocessable",
    "Avg Barcode Per Stop",
    "Avg Barcode Travelled",
    "Avg Bot Availability",
    "Avg Bot Stops",
    "Avg Bot Uptime",
    "Avg Bots",
    "Avg Bots In Maintenance",
    "Avg Breached Percentage",
    "Avg Charger Availability",
    "Avg Charger Success Rate",
    "Avg Charger Uptime",
    "Avg Chargers On Site",
    "Avg Dead Locations",
    "Avg Items Picked",
    "Avg Locations",
    "Avg MOWT Per Unit",
    "Avg MSU",
    "Avg MSU Location Utilization",
    "Avg Orderline Per Order",
    "Avg Orderpool",
    "Avg Pick Exceptions",
    "Avg Pick R2R",
    "Avg Pick UPH",
    "Avg Picks Per Face",
    "Avg Put Exceptions",
    "Avg Put Per Face",
    "Avg Put UPH",
    "Avg QTY Per Orderline",
    "Avg Replened Units",
    "Avg Resolution Time",
    "Avg SKU",
    "Avg Total QTY",
]


def get_influx_client():
    return InfluxDBClient(
        host='alteryx-influx.greymatter.greyorange.com',
        port=8086,
        database='Alteryx',
        username='altryx',
        password='Z20tcH!vZC$jb3V&wYW$ZGFlZ3U',
        ssl=True
    )


def get_empty_column_from_excel_sheet(wb):
    ws = wb.active
    for a in ws.iter_cols(min_row=2, min_col=0, max_col=99999):
        if a[0].value == None:
            return a[0].column_letter


def create_new_workbook(file_location):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "1st to 7th"
    ws["C1"] = "8th to 14th"
    ws["D1"] = "15th to 21st"
    ws["E1"] = "22nd to remaining"
    ws["F1"] = "Monthly_Avg"
    for index, kpi in enumerate(KPI_NAMES):
        ws[f"A{index+2}"] = kpi
    wb.save(file_location)
    return wb


def append_new_data(wb, data):
    col_letter = get_empty_column_from_excel_sheet(wb)
    ws = wb.active
    for row_index, value in enumerate(list(data.values())):
        ws[f"{col_letter}{row_index+2}"] = value


def get_previous_month(today):
    first = today.replace(day=1)
    return first - timedelta(days=1)


def get_time_ranges():
    today = date.today()
    today = datetime.strptime("2022-12-03", "%Y-%m-%d")
    last_month = get_previous_month(today)
    start_datetime = datetime.strptime("{}-{}-01 05:00:00".format(
        last_month.year, last_month.month), "%Y-%m-%d %H:%M:%S")
    end_datetime = datetime.strptime("{}-{}-01 05:00:00".format(
        today.year, today.month), "%Y-%m-%d %H:%M:%S")
    out_format = "%Y-%m-%d %H:%M:%S"
    return [
        {
            "id": "1st_to_7th",
            "name": "1st to 7th",
            "range": [
                datetime.strftime(start_datetime, out_format),
                datetime.strftime(start_datetime + timedelta(days=7), out_format),
            ]
        }, {
            "id": "8th_to_14th",
            "name": "8th to 14th",
            "range": [
                datetime.strftime(start_datetime + timedelta(days=7), out_format),
                datetime.strftime(start_datetime + timedelta(days=14), out_format),
            ]
        }, {
            "id": "15th_to_21st",
            "name": "15th to 21st",
            "range": [
                datetime.strftime(start_datetime + timedelta(days=14), out_format),
                datetime.strftime(start_datetime + timedelta(days=21), out_format),
            ]
        }, {
            "id": "22nd_to_remaining",
            "name": "22nd to remaining",
            "range": [
                datetime.strftime(start_datetime + timedelta(days=21), out_format),
                datetime.strftime(end_datetime, out_format),
            ]
        }, {
            "id": "Monthly_Avg",
            "name": "Monthly_Avg",
            "range": [
                datetime.strftime(start_datetime, out_format),
                datetime.strftime(end_datetime, out_format),
            ]
        },
    ]


def map_influx_result(result, many=False):
    if not many:
        if not result.raw.get("series"):
            return {}

        series = result.raw.get("series")[0]
        return dict(zip(series.get("columns"), series.get("values")[0]))



def get_task_id(_site, _range):
    return f"retrieving_data_{_site.get('id')}__{_range.get('id')}"


def retrieve_data(**kwargs):
    client = get_influx_client()
    _site = kwargs.get('site', {})
    _range = kwargs.get('_range', {})
    result = client.query("""
        select
            mean(audit_completed) as "Avg Audit Completed",
            mean(audit_not_started) as "Avg Audit Not Started",
            mean(audit_unprocessable) as "Avg Audit Unprocessable",
            mean(barcode_per_stop) as "Avg Barcode Per Stop",
            mean(barcode_travelled) as "Avg Barcode Travelled",
            mean(bot_availability) as "Avg Bot Availability",
            mean(bot_stops) as "Avg Bot Stops",
            mean(bot_uptime) as "Avg Bot Uptime",
            mean(total_bots) as "Avg Bots",
            mean(bots_in_maintenance) as "Avg Bots In Maintenance",
            mean(breached_percentage) as "Avg Breached Percentage",
            mean(charger_availability) as "Avg Charger Availability",
            mean(charger_success_rate) as "Avg Charger Success Rate",
            mean(charger_uptime) as "Avg Charger Uptime",
            mean(inducted_chargers) as "Avg Chargers On Site",
            mean(total_dead_locations) as "Avg Dead Locations",
            mean(total_locations) as "Avg Locations",
            (
                sum(SUM_value_mowt_numer) / sum(COUNT_value_mowt_numer)
            ) /(
                sum(SUM_value_mowt_denom) / sum(COUNT_value_mowt_denom)
            ) as "Avg MOWT Per Unit",
            mean(total_msu) as "Avg MSU",
            mean(msu_loc_utilization) as "Avg MSU Location Utilization",
            sum(SUM_line_count) / sum(SUM_order_id_count) as "Avg Orderline Per Order",
            mean(orderpool) as "Avg Orderpool",
            mean(item_exceptions) as "Avg Pick Exceptions",
            sum(SUM_value_r2r) / sum(COUNT_value_r2r) as "Avg Pick R2R",
            sum(SUM_UPH_60min) / sum(COUNT_UPH_60min) as "Avg Pick UPH",
            mean(SUM_value_mowt_denom) as "Avg Picks Per Face",
            mean(item_put_exceptions) as "Avg Put Exceptions",
            mean(put_per_face) as "Avg Put Per Face",
            sum(SUM_put_UPH_60min) / sum(COUNT_put_UPH_60min) as "Avg Put UPH",
            sum(SUM_qty_per_orderline) / sum(COUNT_qty_per_orderline) as "Avg QTY Per Orderline",
            mean(items_put) as "Avg Replened Units",
            sum(SUM_time_difference) / sum(COUNT_time_difference) as "Avg Resolution Time",
            mean(total_sku) as "Avg SKU",
            mean(total_qty) as "Avg Total QTY"
        from GXO_monthly_report
        where
            time >= '{}'
            and time < '{}'
            and site = '{}'
    """.format(
        *[*_range.get('range'), _site.get('name')]
    ))
    data = map_influx_result(result)

    avg_item_picked = map_influx_result(
        client.query("""
            select
                mean(items_picked) as Avg_items_picked
            from
                GXO_monthly_report
            where
                time >= '{}'
                and time < '{}'
                and site = '{}'
                and items_picked > 0
        """.format(*[*_range.get('range'), _site.get('name')])
        )
    )
    data["Avg Items Picked"] = avg_item_picked.get("Avg_items_picked")
    final_result = OrderedDict()

    for kpi in KPI_NAMES:
        if data.get(kpi) is not None:
            final_result[kpi] = round(data.get(kpi), 1)

    return final_result


def create_excel_file(**kwargs):
    task_instance = kwargs["task_instance"]
    _site = kwargs.get('site', {})
    result = {}
    file_location = _site.get('file')
    wb = create_new_workbook(file_location)

    for _range in get_time_ranges():
        data = task_instance.xcom_pull(task_ids=get_task_id(_site, _range))
        append_new_data(wb, data)
        result[_range.get('name')] = data
    
    wb.save(file_location)
    return result


def send_mail(**kwargs):
    task_instance = kwargs["task_instance"]
    month_year = datetime.strftime(
        get_previous_month(datetime.today()),
        "%B, %Y"
    )
    environment = Environment()
    template = environment.from_string(MAIL_TEMPLATE)
    template_data = []

    for _site in SITES:
        data = task_instance.xcom_pull(task_ids=f"creating_excel_file_{site.get('id')}")
        data_values = list(data.values())
        template_data.append({
            "name": _site.get('name'),
            "rows": [{
                "name": KPI_NAMES[i],
                "data_list": [
                    list(data_values[0].values())[i],
                    list(data_values[1].values())[i],
                    list(data_values[2].values())[i],
                    list(data_values[3].values())[i],
                    list(data_values[4].values())[i],
                ]
            } for i in range(len(KPI_NAMES))] 
        })
    print(json.dumps(template_data, indent=4))
    message = template.render(
        sites_data_list=template_data
    )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"GXO monthly report ({month_year})"
    msg["From"] = SODIMAC_SENDER_EMAIL
    msg["To"] = SODIMAC_RECEIVER_EMAIL

    msg.attach(MIMEText(message, "html"))

    for _site in SITES:
        file_part = MIMEBase("application", "octet-stream")
        file_part.set_payload(open(_site.get('file'), "rb").read())
        encoders.encode_base64(file_part)
        file_part.add_header(
            "Content-Disposition", 'attachment; filename="gxo_monthly_report.xlsx"'
        )
        msg.attach(file_part)

    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, 587) as server:
        server.starttls(context=context)
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        
        server.sendmail(
            SODIMAC_SENDER_EMAIL,
            SODIMAC_RECEIVER_EMAIL,
            msg.as_string()
        )




with DAG(
    "GXO_monthly_report",
    start_date=airflow.utils.dates.days_ago(0),
    schedule_interval="0 0 */2 * *",
    default_args=CommonFunction().get_default_args_for_dag(),
    catchup=False,
) as dag:

    sending_mail = PythonOperator(
        task_id="sending_mail", python_callable=send_mail)

    for site in SITES:
        creating_excel_file = PythonOperator(
            task_id=f"creating_excel_file_{site.get('id')}",
            python_callable=create_excel_file,
            op_kwargs={'site': site}
        )

        for _range in get_time_ranges():
            retrieving_data_task = PythonOperator(
                task_id=get_task_id(site, _range),
                python_callable=retrieve_data,
                op_kwargs={'site': site, '_range': _range},
            )
        
            retrieving_data_task >> creating_excel_file

        creating_excel_file >> sending_mail