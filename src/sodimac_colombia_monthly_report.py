import math
import smtplib
import ssl
from collections import OrderedDict
from datetime import date, datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import airflow
import pandas as pd
from airflow import DAG
from jinja2 import Environment
from openpyxl import Workbook, load_workbook

from utils.CommonFunction import SingleTenantDAGBase, MultiTenantDAG, CommonFunction
from config import (
    INFLUX_DATETIME_FORMAT,
    SMTP_USERNAME,
    SMTP_PASSWORD,
    SMTP_HOST,
    SODIMAC_SENDER_EMAIL,
    SODIMAC_RECEIVER_EMAIL,
)


MAIL_TEMPLATE = """
<table style="width: 800px; font-size: 18px; border: solid #f8f9fa;">
    <thead style="background: antiquewhite;">
        <th style="text-align: left;">Name</th>
        <th style="text-align: center;">Value</th>
    </thead>
    <tbody style="background: floralwhite;">
        {% for item in items %}
            <tr>
                <td style="text-align: left;">{{ item.name }}</td>
                <td style="text-align: center;">{{ item.value }}</td>
            </tr>
        {% endfor %}
    </tbody>
</table>
"""

EXCEL_FILE_LOCATION = "dags/files/sodimac_colombia_monthly_report.xlsx"

dag = DAG(
    "sodimac_colombia_monthly_report",
    start_date=airflow.utils.dates.days_ago(0),
    schedule_interval="0 0 */2 * *",
    default_args=CommonFunction().get_default_args_for_dag(),
    catchup=False,
)

######################## Helpers ########################

def get_empty_column_from_excel_sheet(wb):
    ws = wb.active
    for a in ws.iter_cols(min_row=1, min_col=0, max_col=999999):
        if a[0].value == None:
            return a[0].column_letter


######################## TASKS ########################

class SodimacColombiaMonthlyReport(SingleTenantDAGBase):
    def get_previous_month(self, today):
        first = today.replace(day=1)
        return first - timedelta(days=1)

    def get_fetch_time_interval(self):
        # today = date.today()
        today = datetime.strptime('2022-11-05', '%Y-%m-%d')
        last_month = self.get_previous_month(today)
        start_datetime = "{}-{}-01 05:00:00".format(
            last_month.year, str(last_month.month).zfill(2))
        end_datetime = "{}-{}-01 05:00:00".format(today.year, str(today.month).zfill(2))
        return start_datetime, end_datetime

    def fetch_data(self, query):
        client = self.get_site_read_influx_client('GreyOrange')
        return client.fetch_one(
            query.format(*self.get_fetch_time_interval()), format="json"
        )

    def get_qty_picked_aggregated_data(self):
        return self.fetch_data("""
            select
                sum(value) as qty_picked,
                sum(uom_quantity_int) as qty_picked_InnerandEaches,
                sum(volume_cm3) as volume_utilization
            from item_picked
            where time >= '{}' and time < '{}'
        """)

    def get_total_completed_order(self):
        return self.fetch_data("""
            select count(distinct order_id) as total_completed_order
            from order_events
            where
                time >= '{}' and
                time < '{}' and
                event_name = 'order_complete'
        """)

    def get_racks_presented(self):
        return self.fetch_data("""
            select count(value) as racks_presented
            from pps_events
            where
                process_mode = 'pick' and
                time >= '{}' and
                time < '{}'
        """)

    def get_interval_throughput_aggragated_data(self):
        return self.fetch_data("""
            select
                mean(UPH_60min) as picks_per_pps_per_hour,
                mean(UPH_60min_value) as units_per_pps_per_hour
            from interval_throughput
            where time >= '{}' and time < '{}'
        """)

    def get_total_items_from_slots_ageing_ayx(self, ):
        client = self.get_site_read_influx_client('GreyOrange')
        cron_ran_at = client.fetch_one("""
            select cron_ran_at
            from slots_ageing_ayx
            where time >= '{}' and time < '{}'
            order by time desc limit 1
        """.format(*self.get_fetch_time_interval()), format="json").get(
            "cron_ran_at"
        )
        # TODO: This influx query is very slow. need to fix it
        # return map_influx_result(client.query("""
        #     select count(distinct item_id) as sku
        #     from slots_ageing_ayx
        #     where cron_ran_at = '{}'
        # """.format(cron_ran_at)))
        return {
            "sku": 0
        }

    def get_percetile95_qyt_picked(self, ):
        influx = self.get_site_read_influx_client('GreyOrange')
        time_interval = self.get_fetch_time_interval()
        print("time_interval", time_interval)

        item_picked = influx.client.query("""
            select sum(value)
            from item_picked
            where time >= '{}' and time < '{}'
            group by pps_id
        """.format(*time_interval))

        df = pd.DataFrame(
            columns=[
                "time",
                "pps_id",
                "value",
                "uom_int",
                "uom_quantity_int",
                "uom_quantity",
            ]
        )

        for item_picked_series in list(item_picked):
            pps_id = dict(item_picked_series[1]).get("pps_id")

            if not pps_id:
                continue

            item_picked_per_pps_df = influx.fetch_all("""
                select pps_id, value, uom_int, uom_quantity_int, uom_quantity
                from item_picked
                where time >= '{}' and time < '{}' and pps_id = '{}'
            """.format(
                    *[*time_interval, pps_id]
                )
            )

            df = df.append(item_picked_per_pps_df)

        df['time'] = df.index
        df["time"] = df["time"].map(
            lambda t: t.replace(minute=0, second=0).strftime(INFLUX_DATETIME_FORMAT)
        )
        print("df", df)
        df["uom_int"] = df.apply(
            lambda row: row["uom_quantity_int"]
            if row["uom_int"] is None
            else row["uom_int"],
            axis=1,
        )
        df["uom_int"] = df.apply(
            lambda row: row["uom_quantity"]
            if row["uom_int"] is None
            else row["uom_int"],
            axis=1,
        )
        df = df[["time", "pps_id", "value", "uom_int"]
                ].groupby(["time", "pps_id"]).sum()

        high_percentile95_index = math.ceil(len(df) * 0.975)
        low_percentile95_index = math.ceil(len(df) * 0.025)
        percentile_value_up = df.sort_values(
            by=["value"], axis=0
        ).iloc[high_percentile95_index]["value"]
        percentile_value_low = df.sort_values(
            by=["value"], axis=0).iloc[low_percentile95_index]["value"]
        percentile_uom_int_up = df.sort_values(
            by=["uom_int"], axis=0).iloc[high_percentile95_index]["uom_int"]
        percentile_uom_int_low = df.sort_values(
            by=["uom_int"], axis=0).iloc[low_percentile95_index]["uom_int"]

        df["value_flag"] = df["value"].map(
            lambda v: True
            if v >= percentile_value_low and v <= percentile_value_up
            else False
        )
        df["uom_int_flag"] = df["uom_int"].map(
            lambda v: True
            if v >= percentile_uom_int_low and v <= percentile_uom_int_up
            else False
        )
        return {
            "qty_picked_95percentile": float(
                df[df["value_flag"] == True][["value"]].sum()["value"]
            ),
            "qty_picked_InnerAndEaches_95percentile": float(
                df[df["uom_int_flag"] == True][["uom_int"]].sum()["uom_int"]
            ),
        }


    def get_transaction_per_pps_per_hour(self):
        client = self.get_site_read_influx_client('GreyOrange')
        item_picked = client.fetch_all("""
            select count(value)
            from item_picked
            where
                time >= '{}' and
                time < '{}'
            group by time(1h), pps_id
        """.format(*self.get_fetch_time_interval()), format="json")

        df = pd.DataFrame(columns=["time", "count"])

        for item_picked_series in item_picked:
            df = df.append(
                [
                    pd.DataFrame(
                        item_picked_series.get("values"),
                        columns=item_picked_series.get("columns"),
                    )
                ],
                ignore_index=True,
            )

        return float(round(df[df["count"] != 0][["count"]].mean()["count"], 2))


    def get_pps_hr(self):
        return self.fetch_data("""
            select sum(value)/3600 as pps_hr
            from pps_operator_log
            where
                time >= '{}' and
                time < '{}' and
                pps_mode='pick_mode' and
                (
                    type='working_time' or
                    type='waiting_time'
                )
        """)

    def merge_data(self, *args, **kwargs):
        # Getting data from previous tasks
        qty_picked_data = self.get_task_data("get_qty_picked_aggregated_data", **kwargs)
        total_completed_order_data = self.get_task_data("get_total_completed_order", **kwargs)
        racks_presented_data = self.get_task_data("get_racks_presented", **kwargs)
        interval_throughput_aggragated_data = self.get_task_data(
            "get_interval_throughput_aggragated_data", **kwargs
        )
        total_items_from_slots_ageing_ayx = self.get_task_data(
            "get_total_items_from_slots_ageing_ayx", **kwargs
        )
        percetile95_qyt_picked = self.get_task_data("get_percetile95_qyt_picked", **kwargs)
        transaction_per_pps_per_hour = self.get_task_data(
            "get_transaction_per_pps_per_hour", **kwargs
        )
        pps_hr = self.get_task_data("get_pps_hr", **kwargs)

        # Fetching final values
        units = qty_picked_data.get("qty_picked")
        picks = qty_picked_data.get("qty_picked_InnerandEaches")
        orders = total_completed_order_data.get("total_completed_order")
        volume_utilization = qty_picked_data.get("volume_utilization")
        racks_presented = racks_presented_data.get("racks_presented")
        units_per_pps_per_hr = interval_throughput_aggragated_data.get(
            "units_per_pps_per_hour"
        )
        picks_per_pps_per_hour = interval_throughput_aggragated_data.get(
            "picks_per_pps_per_hour"
        )
        sku = total_items_from_slots_ageing_ayx.get("sku")
        units_percentile_95 = percetile95_qyt_picked.get("qty_picked_95percentile")
        picks_percentile_95 = percetile95_qyt_picked.get(
            "qty_picked_InnerAndEaches_95percentile"
        )
        pps_per_hr = pps_hr.get("pps_hr")

        # Calculating final values
        unit_per_face = units / racks_presented
        pick_per_face = picks / racks_presented
        volume_utilization = volume_utilization / 1000000
        volume_utilization_percentage = (volume_utilization / 1150) * 100
        units_per_pick = units / picks
        picks_per_order = picks / orders

        return OrderedDict(
            {
                "Unidades": round(units, 0),
                "Picks": picks,
                "Ordenes (oLPN)": orders,
                "Rack presentations": racks_presented,
                "Units/face": round(unit_per_face, 2),
                "Picks/face": round(pick_per_face, 2),
                "Relacion Unidades Pick": round(units_per_pick, 2),
                "Relacion Unidades OLPN": round(picks_per_order, 2),
                "Units/PPS/hr": round(units_per_pps_per_hr, 2),
                "Picks/PPS/hr": round(picks_per_pps_per_hour, 2),
                "SKUs": sku,
                "Utilizacion del volumen (m3)": round(volume_utilization, 2),
                "Utilizacion del volumen (%)": round(volume_utilization_percentage, 2),
                "Picks (percentile 95)": picks_percentile_95,
                "Unidades (percentile 95)": units_percentile_95,
                "Transaction/PPS/hr": transaction_per_pps_per_hour,
                "PPS/Hr": round(pps_per_hr, 2),
            }
        )


    def create_new_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Mes"
        ws["A2"] = "Año"
        ws["A3"] = "Unidades"
        ws["A4"] = "Picks"
        ws["A5"] = "Ordenes (oLPN)"
        ws["A6"] = "Rack presentations"
        ws["A7"] = "Units/face"
        ws["A8"] = "Picks/face"
        ws["A9"] = "Relación Unidades Pick"
        ws["A10"] = "Relación Unidades OLPN"
        ws["A11"] = "Units/PPS/hr"
        ws["A12"] = "Picks/PPS/hr"
        ws["A13"] = "SKUs"
        ws["A14"] = "Utilizacion del volumen (m3)"
        ws["A15"] = "Utilizacion del volumen (%)"
        ws["A16"] = "Picks (percentile 95)"
        ws["A17"] = "Unidades (percentile 95)"
        ws["A18"] = "Transaction/PPS/hr"
        ws["A19"] = "PPS/Hr"
        wb.save(EXCEL_FILE_LOCATION)
        return wb


    def append_new_data(self, wb, data):
        col_letter = get_empty_column_from_excel_sheet(wb)
        ws = wb.active
        today = datetime.today()
        values = [today.month, today.year] + list(data.values())
        for row_index, value in enumerate(values):
            ws[f"{col_letter}{row_index+1}"] = value

        wb.save(EXCEL_FILE_LOCATION)


    def create_excel_file(self, *args, **kwargs):
        data = self.get_task_data("merge_data", **kwargs)
        try:
            wb = load_workbook(EXCEL_FILE_LOCATION)
        except:
            wb = self.create_new_workbook()

        self.append_new_data(wb, data)


    def send_mail(self, *args, **kwargs):
        data = self.get_task_data("merge_data", **kwargs)

        month_year = datetime.strftime(
            self.get_previous_month(datetime.today()),
            "%B, %Y"
        )

        environment = Environment()
        template = environment.from_string(MAIL_TEMPLATE)
        message = template.render(
            items=[{"name": key, "value": data.get(
                key, "NA")} for key in data.keys()]
        )

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Sodimac Colombia monthly report ({month_year})"
        msg["From"] = SODIMAC_SENDER_EMAIL
        msg["To"] = SODIMAC_RECEIVER_EMAIL

        msg.attach(MIMEText(message, "html"))

        file_part = MIMEBase("application", "octet-stream")
        file_part.set_payload(open(EXCEL_FILE_LOCATION, "rb").read())
        encoders.encode_base64(file_part)
        file_part.add_header(
            "Content-Disposition", 'attachment; filename="sodimac_colombia_report.xlsx"'
        )
        msg.attach(file_part)

        print("SMTP_HOST", SMTP_HOST)
        print("SMTP_USERNAME", SMTP_USERNAME)
        print("SMTP_PASSWORD", SMTP_PASSWORD)

        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_HOST, 587) as server:
            server.starttls(context=context)
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            
            server.sendmail(
                SODIMAC_SENDER_EMAIL,
                SODIMAC_RECEIVER_EMAIL,
                msg.as_string()
            )


######################## DAG ########################

MultiTenantDAG(
    dag,
    [
        'get_qty_picked_aggregated_data',
        'get_total_completed_order',
        'get_racks_presented',
        'get_interval_throughput_aggragated_data',
        'get_total_items_from_slots_ageing_ayx',
        'get_percetile95_qyt_picked',
        'get_transaction_per_pps_per_hour',
        'get_pps_hr',
        'merge_data',
        'send_mail',
        'create_excel_file',
    ],
    [
        'get_qty_picked_aggregated_data >> merge_data',
        'get_total_completed_order >> merge_data',
        'get_racks_presented >> merge_data',
        'get_interval_throughput_aggragated_data >> merge_data',
        'get_total_items_from_slots_ageing_ayx >> merge_data',
        'get_percetile95_qyt_picked >> merge_data',
        'get_transaction_per_pps_per_hour >> merge_data',
        'get_pps_hr >> merge_data',
        'merge_data >> create_excel_file',
        'create_excel_file >> send_mail',
    ],
    SodimacColombiaMonthlyReport
).create()
